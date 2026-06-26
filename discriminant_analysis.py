import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import Workbook, load_workbook

from tk_env import configure_tk_environment

configure_tk_environment()

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

SUPPORTED_INPUT_EXTENSIONS = {".csv", ".xlsx"}
LABEL_COLUMN_NAMES = {"cluster", "label", "target"}


class StandardScalerModel:
    def __init__(self) -> None:
        self.mean_ = None
        self.var_ = None
        self.scale_ = None

    def fit(self, x_values: np.ndarray) -> "StandardScalerModel":
        self.mean_ = np.mean(x_values, axis=0)
        self.var_ = np.var(x_values, axis=0)
        self.scale_ = np.sqrt(self.var_)
        self.scale_[self.scale_ == 0] = 1.0
        return self

    def transform(self, x_values: np.ndarray) -> np.ndarray:
        if self.mean_ is None or self.scale_ is None:
            raise ValueError("標準化モデルが未学習です。")
        return (x_values - self.mean_) / self.scale_

    def fit_transform(self, x_values: np.ndarray) -> np.ndarray:
        return self.fit(x_values).transform(x_values)


class LinearDiscriminantAnalysisModel:
    def __init__(self) -> None:
        self.classes_ = None
        self.priors_ = None
        self.coef_ = None
        self.intercept_ = None

    def fit(self, x_values: np.ndarray, y_values: np.ndarray) -> "LinearDiscriminantAnalysisModel":
        classes, inverse = np.unique(y_values, return_inverse=True)
        n_samples, n_features = x_values.shape
        n_classes = classes.size

        if n_classes < 2:
            raise ValueError("ラベル列には 2 つ以上のグループが必要です。")

        priors = np.zeros(n_classes, dtype=float)
        means = np.zeros((n_classes, n_features), dtype=float)
        pooled_covariance = np.zeros((n_features, n_features), dtype=float)

        for class_index, class_label in enumerate(classes):
            class_mask = inverse == class_index
            class_values = x_values[class_mask]
            class_count = class_values.shape[0]
            priors[class_index] = class_count / n_samples
            means[class_index] = np.mean(class_values, axis=0)

            centered = class_values - means[class_index]
            pooled_covariance += centered.T @ centered

        denominator = n_samples - n_classes
        if denominator <= 0:
            raise ValueError("学習データ数が不足しているため判別分析を実行できません。")

        pooled_covariance /= denominator
        pooled_covariance += np.eye(n_features) * 1e-9
        covariance_inverse = np.linalg.pinv(pooled_covariance)

        coef = means @ covariance_inverse
        intercept = np.empty(n_classes, dtype=float)
        for class_index in range(n_classes):
            intercept[class_index] = (
                -0.5 * means[class_index] @ covariance_inverse @ means[class_index]
                + np.log(priors[class_index])
            )

        self.classes_ = classes
        self.priors_ = priors
        self.coef_ = coef
        self.intercept_ = intercept
        return self

    def decision_function(self, x_values: np.ndarray) -> np.ndarray:
        if self.coef_ is None or self.intercept_ is None:
            raise ValueError("判別分析モデルが未学習です。")
        return x_values @ self.coef_.T + self.intercept_

    def predict(self, x_values: np.ndarray) -> np.ndarray:
        scores = self.decision_function(x_values)
        best_index = np.argmax(scores, axis=1)
        return self.classes_[best_index]

    def score(self, x_values: np.ndarray, y_values: np.ndarray) -> float:
        predictions = self.predict(x_values)
        return float(np.mean(predictions == y_values))


def validate_input_file(file_path: Path) -> Path:
    resolved = Path(file_path).expanduser()
    if not resolved.exists():
        raise FileNotFoundError(f"{resolved} が見つかりません。")
    if resolved.suffix.lower() not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(
            f"{resolved.name} は未対応形式です。対応形式: {', '.join(sorted(SUPPORTED_INPUT_EXTENSIONS))}"
        )
    return resolved


def convert_cell_value(value):
    return "" if value is None else value


def is_header_row(row) -> bool:
    non_empty = [cell for cell in row if str(cell).strip()]
    if len(non_empty) < 2:
        return False
    return isinstance(row[0], str) and bool(str(row[0]).strip())


def normalize_table_rows(rows, file_name: str):
    start_index = 0
    for index, row in enumerate(rows):
        if is_header_row(row):
            start_index = index
            break
    else:
        raise ValueError(f"{file_name} のヘッダーが読み取れません。")

    normalized_rows = rows[start_index:]
    header = normalized_rows[0]
    data_rows = normalized_rows[1:]
    return header, data_rows


def load_csv_table(file_path: Path):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = [[convert_cell_value(cell) for cell in row] for row in reader]

    if not rows:
        raise ValueError(f"{file_path.name} が空です。")

    return normalize_table_rows(rows, file_path.name)


def load_excel_table(file_path: Path):
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    rows = [[convert_cell_value(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]

    if not rows:
        raise ValueError(f"{file_path.name} が空です。")

    return normalize_table_rows(rows, file_path.name)


def load_table(file_path: Path):
    file_path = validate_input_file(file_path)
    if file_path.suffix.lower() == ".csv":
        return load_csv_table(file_path)
    return load_excel_table(file_path)


def coerce_feature_matrix(rows, start_col: int, end_col: Optional[int] = None) -> np.ndarray:
    sliced_rows = [row[start_col:end_col] for row in rows]
    if not sliced_rows:
        raise ValueError("データ行がありません。")

    try:
        return np.asarray(sliced_rows, dtype=float)
    except ValueError as error:
        raise ValueError("特徴量列には数値を入力してください。") from error


def coerce_label_vector(values) -> np.ndarray:
    try:
        return np.asarray(values, dtype=int)
    except ValueError:
        pass

    try:
        return np.asarray(values, dtype=float)
    except ValueError:
        pass

    return np.asarray([str(value) for value in values], dtype=object)


def split_train_features_label(header, rows):
    if len(header) < 3:
        raise ValueError("学習データは ID 列、特徴量列、ラベル列の最低 3 列が必要です。")
    if not rows:
        raise ValueError("学習データに行がありません。")

    ids = [row[0] if row else "" for row in rows]
    x_values = coerce_feature_matrix(rows, 1, -1)
    y_values = coerce_label_vector([row[-1] for row in rows])
    return ids, x_values, y_values


def split_test_features(header, rows):
    if len(header) < 2:
        raise ValueError("テストデータは ID 列と特徴量列の最低 2 列が必要です。")
    if not rows:
        raise ValueError("テストデータに行がありません。")

    ids = [row[0] if row else "" for row in rows]
    last_col_name = str(header[-1]).strip().lower()

    if last_col_name in LABEL_COLUMN_NAMES:
        x_values = coerce_feature_matrix(rows, 1, -1)
        y_values = coerce_label_vector([row[-1] for row in rows])
    else:
        x_values = coerce_feature_matrix(rows, 1, None)
        y_values = None

    return ids, x_values, y_values


def ensure_valid(y_values):
    unique = np.unique(y_values)
    if unique.size < 2:
        raise ValueError("ラベル列には 2 つ以上のグループが必要です。")
    return unique


def timestamp_suffix() -> str:
    return datetime.now().strftime("%m%d%H%M")


def add_timestamp_to_path(base_path: Path, stamp: Optional[str] = None) -> Path:
    stamp = stamp or timestamp_suffix()
    parent = base_path.parent if base_path.parent != Path("") else Path.cwd()
    return parent / f"{base_path.stem}_{stamp}{base_path.suffix}"


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{counter:02d}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def default_output_paths(test_path: Path, output_dir: Optional[Path] = None):
    output_dir = Path(output_dir) if output_dir else test_path.parent
    test_extension = test_path.suffix.lower() if test_path.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS else ".csv"
    report_path = ensure_unique_path(add_timestamp_to_path(output_dir / f"lda_report{test_extension}"))
    test_output_path = ensure_unique_path(
        add_timestamp_to_path(output_dir / f"{test_path.stem}_with_prediction{test_extension}")
    )
    return report_path, test_output_path


def write_csv_table(output_path: Path, header, rows) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def write_excel_table(output_path: Path, header, rows) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(list(header))
    for row in rows:
        worksheet.append(list(row))
    workbook.save(output_path)


def save_table(output_path: Path, header, rows) -> None:
    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        write_csv_table(output_path, header, rows)
        return
    if suffix == ".xlsx":
        write_excel_table(output_path, header, rows)
        return
    raise ValueError(f"{output_path.name} の出力形式は未対応です。")


def _fmt(value):
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.integer):
        return int(value)
    return value


def build_report_rows(
    train_accuracy: float,
    lda: LinearDiscriminantAnalysisModel,
    scaler: StandardScalerModel,
    y_train: np.ndarray,
    train_predictions: np.ndarray,
    test_predictions: np.ndarray,
):
    n_features = lda.coef_.shape[1]
    n_cols = max(n_features + 1, len(lda.classes_) + 2, 4)

    def pad(row):
        row = list(row)
        return row + [""] * (n_cols - len(row))

    empty = pad([])
    header = pad(["item", "value"])

    rows = []
    rows.append(pad(["学習判別率", _fmt(train_accuracy)]))
    rows.append(empty)
    rows.append(pad(["学習データクロス集計", "予測グループ"]))
    rows.append(pad(["クラスター", *list(lda.classes_), "合計"]))
    column_totals = np.zeros(len(lda.classes_), dtype=int)
    for true_label in lda.classes_:
        counts = []
        true_mask = y_train == true_label
        for pred_index, predicted_label in enumerate(lda.classes_):
            count = int(np.sum(true_mask & (train_predictions == predicted_label)))
            counts.append(count)
            column_totals[pred_index] += count
        rows.append(pad([true_label, *counts, int(np.sum(counts))]))
    rows.append(pad(["合計", *[int(total) for total in column_totals], int(np.sum(column_totals))]))
    rows.append(empty)
    rows.append(pad(["標準化", "平均=0, 分散=1 (自前 StandardScaler 適用済み)"]))
    rows.append(empty)
    rows.append(pad(["クラスラベル", *list(lda.classes_)]))
    rows.append(pad(["クラス事前確率", *[_fmt(p) for p in lda.priors_]]))
    rows.append(empty)
    rows.append(["クラス", *[f"coef_{index + 1}" for index in range(n_features)]])
    for index, class_label in enumerate(lda.classes_):
        rows.append([f"クラス {class_label}", *[_fmt(c) for c in lda.coef_[index]]])
    rows.append(empty)
    rows.append(pad(["クラス", "intercept"]))
    for index, class_label in enumerate(lda.classes_):
        rows.append(pad([f"クラス {class_label}", _fmt(lda.intercept_[index])]))
    rows.append(empty)
    rows.append(["標準化平均", *[f"mean_{index + 1}" for index in range(n_features)]])
    rows.append(["値", *[_fmt(m) for m in scaler.mean_]])
    rows.append(empty)
    rows.append(["標準化分散", *[f"var_{index + 1}" for index in range(n_features)]])
    rows.append(["値", *[_fmt(v) for v in scaler.var_]])
    rows.append(empty)
    rows.append(pad(["テストデータ予測グループ集計分布"]))
    rows.append(pad(["predicted_label", "count", "ratio"]))
    test_total = len(test_predictions)
    for class_label in lda.classes_:
        count = int(np.sum(test_predictions == class_label))
        ratio = count / test_total if test_total else 0
        rows.append(pad([class_label, count, _fmt(ratio)]))
    rows.append(pad(["total", test_total, _fmt(1.0 if test_total else 0)]))

    return header, rows


def append_prediction_column(header, rows, predictions):
    output_header = list(header) + ["predicted_label"]
    output_rows = [list(row) + [prediction] for row, prediction in zip(rows, predictions)]
    return output_header, output_rows


def run_analysis(
    train_path: Path,
    test_path: Path,
    output_dir: Optional[Path] = None,
    output_report: Optional[Path] = None,
    output_test: Optional[Path] = None,
):
    train_header, train_rows = load_table(train_path)
    test_header, test_rows = load_table(test_path)

    _, x_train, y_train = split_train_features_label(train_header, train_rows)
    _, x_test, _ = split_test_features(test_header, test_rows)

    ensure_valid(y_train)

    if x_train.shape[1] != x_test.shape[1]:
        raise ValueError(
            f"学習データの特徴量数 ({x_train.shape[1]}) とテストデータの特徴量数 ({x_test.shape[1]}) が一致しません。"
        )

    if output_dir:
        output_dir = Path(output_dir).expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)

    default_report_path, default_test_output_path = default_output_paths(Path(test_path), output_dir)
    report_path = ensure_unique_path(add_timestamp_to_path(Path(output_report))) if output_report else default_report_path
    test_output_path = ensure_unique_path(add_timestamp_to_path(Path(output_test))) if output_test else default_test_output_path
    report_path.parent.mkdir(parents=True, exist_ok=True)
    test_output_path.parent.mkdir(parents=True, exist_ok=True)

    scaler = StandardScalerModel()
    x_train_std = scaler.fit_transform(x_train)
    x_test_std = scaler.transform(x_test)

    lda = LinearDiscriminantAnalysisModel()
    lda.fit(x_train_std, y_train)

    y_pred = lda.predict(x_test_std)
    train_predictions = lda.predict(x_train_std)
    train_accuracy = float(np.mean(train_predictions == y_train))
    report_header, report_rows = build_report_rows(
        train_accuracy, lda, scaler, y_train, train_predictions, y_pred
    )
    save_table(report_path, report_header, report_rows)

    test_output_header, test_output_rows = append_prediction_column(test_header, test_rows, y_pred)
    save_table(test_output_path, test_output_header, test_output_rows)

    return {
        "train_accuracy": train_accuracy,
        "report_path": report_path,
        "test_output_path": test_output_path,
    }


class DiscriminantAnalysisApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("判別分析 (LDA)")
        self.root.geometry("760x280")
        self.root.minsize(720, 260)

        self.train_var = tk.StringVar()
        self.test_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.status_var = tk.StringVar(value="学習用ファイルとテスト用ファイルを選択してください。")

        self._build_widgets()

    def _build_widgets(self) -> None:
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill="both", expand=True)

        self.root.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="学習ファイル").grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.train_var).grid(row=0, column=1, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(frame, text="参照", command=self.select_train_file).grid(row=0, column=2, pady=(0, 8))

        ttk.Label(frame, text="テストファイル").grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.test_var).grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(frame, text="参照", command=self.select_test_file).grid(row=1, column=2, pady=(0, 8))

        ttk.Label(frame, text="出力フォルダ").grid(row=2, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(frame, textvariable=self.output_dir_var).grid(row=2, column=1, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(frame, text="参照", command=self.select_output_dir).grid(row=2, column=2, pady=(0, 8))

        note = (
            "対応形式: csv / xlsx\n"
            "出力ファイル名には月日時分のタイムスタンプを自動付与し、既存ファイルを上書きしません。"
        )
        ttk.Label(frame, text=note, justify="left").grid(row=3, column=0, columnspan=3, sticky="w", pady=(6, 16))

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=4, column=0, columnspan=3, sticky="ew")
        button_frame.columnconfigure(0, weight=1)

        self.run_button = ttk.Button(button_frame, text="実行", command=self.run)
        self.run_button.grid(row=0, column=0, sticky="e")

        ttk.Separator(frame).grid(row=5, column=0, columnspan=3, sticky="ew", pady=12)
        ttk.Label(frame, textvariable=self.status_var, justify="left").grid(row=6, column=0, columnspan=3, sticky="w")

    def select_train_file(self) -> None:
        path = self._ask_open_file()
        if path:
            self.train_var.set(path)
            self._fill_output_dir_from_input(path)

    def select_test_file(self) -> None:
        path = self._ask_open_file()
        if path:
            self.test_var.set(path)
            self._fill_output_dir_from_input(path)

    def select_output_dir(self) -> None:
        selected = filedialog.askdirectory(
            title="出力フォルダを選択",
            initialdir=self.output_dir_var.get() or str(Path.cwd()),
        )
        if selected:
            self.output_dir_var.set(selected)

    def _fill_output_dir_from_input(self, selected_path: str) -> None:
        if not self.output_dir_var.get():
            self.output_dir_var.set(str(Path(selected_path).parent))

    def _ask_open_file(self) -> str:
        filetypes = [
            ("対応ファイル", "*.csv *.xlsx"),
            ("CSV ファイル", "*.csv"),
            ("Excel ファイル", "*.xlsx"),
        ]
        return filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=filetypes,
            initialdir=self.output_dir_var.get() or str(Path.cwd()),
        )

    def run(self) -> None:
        train_path = self.train_var.get().strip()
        test_path = self.test_var.get().strip()
        output_dir = self.output_dir_var.get().strip()

        if not train_path or not test_path:
            messagebox.showerror("入力不足", "学習ファイルとテストファイルを選択してください。")
            return

        self.run_button.state(["disabled"])
        self.status_var.set("処理を実行しています...")
        self.root.update_idletasks()

        try:
            result = run_analysis(
                train_path=Path(train_path),
                test_path=Path(test_path),
                output_dir=Path(output_dir) if output_dir else None,
            )
        except FileNotFoundError as error:
            messagebox.showerror("ファイルエラー", str(error))
            self.status_var.set(str(error))
        except ValueError as error:
            messagebox.showerror("入力エラー", str(error))
            self.status_var.set(str(error))
        except Exception as error:
            messagebox.showerror("予期せぬエラー", str(error))
            self.status_var.set(str(error))
        else:
            summary = (
                f"学習判別率: {result['train_accuracy']:.6f}\n"
                f"レポート: {result['report_path']}\n"
                f"予測結果: {result['test_output_path']}"
            )
            messagebox.showinfo("完了", summary)
            self.status_var.set(summary)
        finally:
            self.run_button.state(["!disabled"])


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="判別分析ツール (LDA)")
    parser.add_argument("--train", type=Path, help="学習用ファイルパス")
    parser.add_argument("--test", type=Path, help="テスト用ファイルパス")
    parser.add_argument("--output-dir", type=Path, help="出力先フォルダ")
    parser.add_argument("--output-report", type=Path, help="レポート出力ファイル")
    parser.add_argument("--output-test", type=Path, help="予測結果出力ファイル")
    parser.add_argument("--gui", action="store_true", help="GUI を起動する")
    return parser


def launch_gui() -> None:
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    DiscriminantAnalysisApp(root)
    root.mainloop()


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.gui or len(sys.argv) == 1:
        launch_gui()
        return

    if not args.train or not args.test:
        parser.error("--train と --test を指定するか、引数なしで GUI を起動してください。")

    try:
        result = run_analysis(
            train_path=args.train,
            test_path=args.test,
            output_dir=args.output_dir,
            output_report=args.output_report,
            output_test=args.output_test,
        )
        print("===== 判別分析完了 =====")
        print(f"学習判別率: {result['train_accuracy']:.6f}")
        print(f"レポート出力: {result['report_path']}")
        print(f"テスト出力: {result['test_output_path']}")
    except FileNotFoundError as error:
        print(f"ファイルが見つかりません: {error}", file=sys.stderr)
        sys.exit(1)
    except ValueError as error:
        print(f"入力データ検証エラー: {error}", file=sys.stderr)
        sys.exit(1)
    except Exception as error:
        print(f"予期せぬエラー: {error}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
